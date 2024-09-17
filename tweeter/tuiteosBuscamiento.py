import pymongo
from pymongo import MongoClient
cMongo = MongoClient('localhost', 27017)
db = cMongo.datos
coleccion = db.ifttttuiteos

planilla = './buscamiento (38).xlsx'
from openpyxl import load_workbook
wb = load_workbook(planilla)
sheet = wb.get_sheet_by_name('Hoja 1')

j = 1
for i in sheet:
    celdaA = 'A' + str(j)           #fecha
    fecha = sheet[celdaA].value

    celdaB = 'B' + str(j)           #usuario
    usuario = sheet[celdaB].value

    celdaC = 'C' + str(j)           #texto
    texto = sheet[celdaC].value

    celdaD = 'D' + str(j)           #url tuit
    url = sheet[celdaD].value
    
    registro = ""
    registro = {"usuario" : usuario,
                "fecha" : fecha,
                "texto" : texto,
                "urlTweet" : url
            }
    print ("Registro       : "  + str(j)  + "\n" \
           "Fecha          : "  + fecha + "\n" \
           "Usuario        : "  + usuario + "\n" \
           "Texto          : "  + texto   + "\n" \
           "Url Tweet      : "  + url   + "\n" )
    #post_id = coleccion.insert_one(registro).inserted_id
    j = j + 1

# Retrieve the value of a certain cell

#print("\n\nFecha: " + sheet['A2'].value + "\n\n" \
#      "Usuario: "  + sheet['B2'].value + "\n\n" \
#      "Texto: " + sheet['C2'].value + "\n\n" \
#      "Url tweet:" + sheet['D2'].value + "\n\n" \
#      "Coords: " + sheet['E2'].value)

# Get currently active sheet
#anotherSheet = wb.active

# Check `anotherSheet` 
#anotherSheet

# Retrieve the value of a certain cell
#sheet['A1'].value

# Select element 'B2' of your sheet 
#c = sheet['B2']

# Retrieve the row number of your element
#c.row

# Retrieve the column letter of your element
#c.column

# Retrieve the coordinates of the cell 
#c.coordinate

# Retrieve cell value 
#sheet.cell(row=1, column=2).value

# Print out values in column 2 
#for i in range(1, 4):
#     print(i, sheet.cell(row=i, column=2).value)
# Import relevant modules from `openpyxl.utils`
#from openpyxl.utils import get_column_letter, column_index_from_string

# Return 'A'
#get_column_letter(1)

# Return '1'
#column_index_from_string('A')

# Print row per row
#for cellObj in sheet['A1':'C3']:
#      for cell in cellObj:
#              print(cells.coordinate, cells.value)
#      print('--- END ---')

# Retrieve the maximum amount of rows 
#sheet.max_row

# Retrieve the maximum amount of columns
#sheet.max_column







































        
    
        
        
        
        
